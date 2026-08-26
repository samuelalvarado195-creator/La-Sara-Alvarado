import flet as ft
import json
import os
import hashlib
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Archivos de persistencia
ARCHIVO_DATOS = "inventario_data.json"
ARCHIVO_EXCEL = "Inventario_Real.xlsx"
ARCHIVO_USUARIOS = "usuarios_data.json"

# ----------------------------------------------------------------------
# FUNCIONES DE SEGURIDAD Y USUARIOS
# ----------------------------------------------------------------------
def hash_pin(pin: str) -> str:
    """Genera un hash SHA-256 del PIN de 6 dígitos por seguridad."""
    return hashlib.sha256(pin.encode("utf-8")).hexdigest()

def cargar_usuarios():
    if os.path.exists(ARCHIVO_USUARIOS):
        with open(ARCHIVO_USUARIOS, "r", encoding="utf-8") as archivo:
            try:
                return json.load(archivo)
            except Exception:
                return []
    return []

def guardar_usuario(nombre, usuario, pin_6_digitos):
    usuarios = cargar_usuarios()
    nuevo_usr = {
        "nombre": nombre,
        "usuario": usuario.lower(),
        "pin_hash": hash_pin(pin_6_digitos),
        "fecha_registro": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    usuarios.append(nuevo_usr)
    with open(ARCHIVO_USUARIOS, "w", encoding="utf-8") as archivo:
        json.dump(usuarios, archivo, indent=4, ensure_ascii=False)

def verificar_pin(pin_ingresado: str) -> bool:
    usuarios = cargar_usuarios()
    pin_h = hash_pin(pin_ingresado)
    for u in usuarios:
        if u.get("pin_hash") == pin_h:
            return True
    return False

# ----------------------------------------------------------------------
# FUNCIONES DE PERSISTENCIA (JSON Y EXCEL)
# ----------------------------------------------------------------------
def cargar_inventario():
    if os.path.exists(ARCHIVO_DATOS):
        with open(ARCHIVO_DATOS, "r", encoding="utf-8") as archivo:
            try:
                return json.load(archivo)
            except Exception:
                return []
    return []

def guardar_inventario(inventario):
    with open(ARCHIVO_DATOS, "w", encoding="utf-8") as archivo:
        json.dump(inventario, archivo, indent=4, ensure_ascii=False)
    actualizar_excel_stock(inventario)

def actualizar_excel_stock(inventario):
    try:
        if os.path.exists(ARCHIVO_EXCEL):
            try:
                wb = load_workbook(ARCHIVO_EXCEL)
            except Exception:
                wb = Workbook()
        else:
            wb = Workbook()

        if "Stock y Finanzas" in wb.sheetnames:
            hoja_stock = wb["Stock y Finanzas"]
            hoja_stock.delete_rows(1, hoja_stock.max_row + 1)
        else:
            hoja_stock = wb.active
            hoja_stock.title = "Stock y Finanzas"

        hoja_stock.append(["Código", "Producto", "Cantidad", "P. Costo (Unidad)", "P. Venta (Unidad)", "Ganancia Total Stock"])
        
        for prod in inventario:
            codigo = prod.get("codigo", "N/A")
            try:
                costo = float(prod.get("precio_costo", 0.0))
                venta = float(prod.get("precio_venta", 0.0))
                cantidad = int(prod.get("cantidad", 0))
                ganancia_total = round((venta - costo) * cantidad, 2)
            except Exception:
                costo, venta, ganancia_total, cantidad = 0.0, 0.0, 0.0, 0
                
            hoja_stock.append([codigo, prod.get("nombre", "Desconocido"), cantidad, costo, venta, ganancia_total])
        
        if "Registro de Ventas" not in wb.sheetnames:
            hoja_ventas = wb.create_sheet(title="Registro de Ventas")
            hoja_ventas.append(["Fecha y Hora", "Código", "Producto", "Cantidad Vendida", "Precio Venta Ud.", "Ganancia Venta"])
            hoja_ventas.append(["TOTALES", "", "", 0, "", 0])

        wb.save(ARCHIVO_EXCEL)
    except Exception as e:
        print(f"Error al guardar Excel: {e}")

def registrar_venta_en_excel(codigo, nombre, cantidad, precio_venta, ganancia_venta):
    try:
        if not os.path.exists(ARCHIVO_EXCEL):
            actualizar_excel_stock([])

        wb = load_workbook(ARCHIVO_EXCEL)
        if "Registro de Ventas" not in wb.sheetnames:
            hoja_ventas = wb.create_sheet(title="Registro de Ventas")
            hoja_ventas.append(["Fecha y Hora", "Código", "Producto", "Cantidad Vendida", "Precio Venta Ud.", "Ganancia Venta"])
        else:
            hoja_ventas = wb["Registro de Ventas"]

        if hoja_ventas.max_row >= 2 and hoja_ventas.cell(row=hoja_ventas.max_row, column=1).value == "TOTALES":
            hoja_ventas.delete_rows(hoja_ventas.max_row, 1)
        
        nueva_fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        hoja_ventas.append([nueva_fecha, codigo, nombre, cantidad, precio_venta, ganancia_venta])
        
        total_cantidades = 0
        total_ganancias = 0
        for fila in range(2, hoja_ventas.max_row + 1):
            total_cantidades += int(hoja_ventas.cell(row=fila, column=4).value or 0)
            total_ganancias += float(hoja_ventas.cell(row=fila, column=6).value or 0)
                
        hoja_ventas.append(["TOTALES", "", "", total_cantidades, "", round(total_ganancias, 2)])
        wb.save(ARCHIVO_EXCEL)
    except Exception as e:
        print(f"Error venta Excel: {e}")

# ----------------------------------------------------------------------
# APLICACIÓN INTERFAZ FLET CON PROTOCOLO DE SEGURIDAD
# ----------------------------------------------------------------------
def main(page: ft.Page):
    page.title = "Control Financiero - Donde La Sarita"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.padding = 10

    inventario = cargar_inventario()
    actualizar_excel_stock(inventario)

    def mostrar_mensaje(msg, color=ft.Colors.GREEN_600):
        page.snack_bar = ft.SnackBar(ft.Text(msg), bgcolor=color)
        page.snack_bar.open = True
        page.update()

    # ------------------------------------------------------------------
    # MÓDULO DE INVENTARIO Y VISTAS DE LA APP
    # ------------------------------------------------------------------
    lista_productos_view = ft.ListView(expand=True, spacing=10)

    def actualizar_vista(lista=None):
        nonlocal inventario
        if lista is None:
            lista = inventario

        lista_productos_view.controls.clear()

        if not lista:
            lista_productos_view.controls.append(
                ft.Container(content=ft.Text("No hay productos registrados.", italic=True, color=ft.Colors.GREY_600), padding=20)
            )
        else:
            for prod in lista:
                codigo = str(prod.get("codigo", "N/A"))
                nombre = prod.get("nombre", "Sin Nombre")
                cant = int(prod.get("cantidad", 0))
                costo = float(prod.get("precio_costo", 0.0))
                venta = float(prod.get("precio_venta", 0.0))
                ganancia = round((venta - costo) * cant, 2)
                color_stock = ft.Colors.RED_500 if cant <= 3 else ft.Colors.GREEN_600

                card_item = ft.Card(
                    elevation=2,
                    content=ft.Container(
                        padding=12,
                        content=ft.Column([
                            ft.Row([
                                ft.Text(nombre, weight=ft.FontWeight.BOLD, size=16, expand=True),
                                ft.Container(
                                    content=ft.Text(f"Stock: {cant}", color=ft.Colors.WHITE, size=12, weight=ft.FontWeight.BOLD),
                                    bgcolor=color_stock, padding=ft.Padding(8, 4, 8, 4), border_radius=12
                                )
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            ft.Text(f"Código: {codigo}", size=12, color=ft.Colors.GREY_700),
                            ft.Row([
                                ft.Text(f"Costo: L. {costo:.2f}", size=12),
                                ft.Text(f"Venta: L. {venta:.2f}", size=12, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
                                ft.Text(f"Ganancia: L. {ganancia:.2f}", size=12, color=ft.Colors.GREEN_700)
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            ft.Divider(height=10, thickness=0.5),
                            ft.Row([
                                ft.IconButton(icon=ft.Icons.SHOPPING_CART_CHECK_OUTLINED, icon_color=ft.Colors.GREEN_600, tooltip="Vender", on_click=lambda e, p=prod: abrir_dialogo_venta(p)),
                                ft.IconButton(icon=ft.Icons.EDIT_OUTLINED, icon_color=ft.Colors.BLUE_600, tooltip="Editar", on_click=lambda e, p=prod: abrir_dialogo_producto(p)),
                                ft.IconButton(icon=ft.Icons.DELETE_OUTLINE, icon_color=ft.Colors.RED_600, tooltip="Eliminar", on_click=lambda e, p=prod: eliminar_producto(p)),
                            ], alignment=ft.MainAxisAlignment.END)
                        ])
                    )
                )
                lista_productos_view.controls.append(card_item)

        page.update()

    def buscar_prod(e):
        termino = txt_buscar.value.strip().lower()
        if not termino:
            actualizar_vista()
            return
        filtrados = [p for p in inventario if termino in str(p.get("codigo", "")).lower() or termino in p.get("nombre", "").lower()]
        actualizar_vista(filtrados)

    # LECTOR DE CÁMARA
    def abrir_escaner_camara(e):
        try:
            import cv2
            from pyzbar.pyzbar import decode
        except ModuleNotFoundError:
            mostrar_mensaje("Ejecuta en consola: pip install opencv-python pyzbar", ft.Colors.RED_500)
            return

        mostrar_mensaje("Abriendo cámara para escáner...", ft.Colors.BLUE_600)
        codigo_detectado = None
        cap = cv2.VideoCapture(0)

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
            for item in decode(frame):
                codigo_detectado = item.data.decode("utf-8")
                break
            if codigo_detectado:
                break

        cap.release()
        cv2.destroyAllWindows()

        if codigo_detectado:
            prod_encontrado = next((p for p in inventario if str(p.get("codigo", "")).lower() == codigo_detectado.lower()), None)
            if prod_encontrado:
                abrir_dialogo_producto(prod_editar=prod_encontrado)
            else:
                abrir_dialogo_producto(codigo_inicial=codigo_detectado)
        else:
            mostrar_mensaje("Escaneo cancelado.", ft.Colors.RED_400)

    txt_buscar = ft.TextField(
        hint_text="Buscar por nombre o código...",
        prefix_icon=ft.Icons.SEARCH,
        suffix=ft.IconButton(icon=ft.Icons.CAMERA_ALT_OUTLINED, icon_color=ft.Colors.BLUE_700, on_click=abrir_escaner_camara),
        on_change=buscar_prod,
        expand=True,
        border_radius=25,
        content_padding=ft.Padding(15, 10, 15, 10)
    )

    # DIÁLOGOS DE EDICIÓN Y CREACIÓN
    def abrir_dialogo_producto(prod_editar=None, codigo_inicial=None):
        codigo_val = str(prod_editar.get("codigo", "")) if prod_editar else (codigo_inicial or "")
        txt_c = ft.TextField(label="Código Único", value=codigo_val)
        txt_n = ft.TextField(label="Nombre del Producto", value=prod_editar.get("nombre", "") if prod_editar else "")
        txt_cant = ft.TextField(label="Cantidad en Stock", value=str(prod_editar.get("cantidad", "")) if prod_editar else "", keyboard_type=ft.KeyboardType.NUMBER)
        txt_cost = ft.TextField(label="Precio Costo Ud. (L.)", value=str(prod_editar.get("precio_costo", "")) if prod_editar else "", keyboard_type=ft.KeyboardType.NUMBER)
        txt_vent = ft.TextField(label="Precio Venta Ud. (L.)", value=str(prod_editar.get("precio_venta", "")) if prod_editar else "", keyboard_type=ft.KeyboardType.NUMBER)

        def guardar(e):
            c, n = txt_c.value.strip(), txt_n.value.strip()
            try:
                cant = int(txt_cant.value.strip())
                costo = float(txt_cost.value.strip().replace('L.', ''))
                venta = float(txt_vent.value.strip().replace('L.', ''))
            except ValueError:
                mostrar_mensaje("Revisa los valores numéricos.", ft.Colors.RED_500)
                return

            if not c or not n:
                mostrar_mensaje("Código y Nombre son obligatorios.", ft.Colors.RED_500)
                return

            if prod_editar:
                prod_editar.update({"codigo": c, "nombre": n, "cantidad": cant, "precio_costo": costo, "precio_venta": venta})
            else:
                existente = next((p for p in inventario if str(p.get("codigo", "")).lower() == c.lower()), None)
                if existente:
                    existente.update({"nombre": n, "cantidad": cant, "precio_costo": costo, "precio_venta": venta})
                else:
                    inventario.append({"codigo": c, "nombre": n, "cantidad": cant, "precio_costo": costo, "precio_venta": venta})

            guardar_inventario(inventario)
            actualizar_vista()
            page.dialog.open = False
            mostrar_mensaje("¡Datos guardados y sincronizados con Excel!")
            page.update()

        page.dialog = ft.AlertDialog(
            title=ft.Text("Editar Producto" if prod_editar else "Nuevo Producto"),
            content=ft.Column([txt_c, txt_n, txt_cant, txt_cost, txt_vent], tight=True, spacing=8),
            actions=[
                ft.ElevatedButton("Guardar en Excel", on_click=guardar, bgcolor=ft.Colors.GREEN_600, color=ft.Colors.WHITE),
                ft.TextButton("Cancelar", on_click=lambda _: setattr(page.dialog, 'open', False) or page.update())
            ]
        )
        page.dialog.open = True
        page.update()

    def abrir_dialogo_venta(prod):
        txt_cant_venta = ft.TextField(label="Cantidad a vender", value="1", keyboard_type=ft.KeyboardType.NUMBER, autofocus=True)

        def confirmar_venta(e):
            try:
                cant_salida = int(txt_cant_venta.value.strip())
            except ValueError:
                mostrar_mensaje("Ingresa una cantidad válida.", ft.Colors.RED_500)
                return

            if cant_salida > int(prod["cantidad"]):
                mostrar_mensaje(f"Stock insuficiente. Disponible: {prod['cantidad']}", ft.Colors.RED_500)
                return

            prod["cantidad"] = int(prod["cantidad"]) - cant_salida
            gan_v = round((float(prod["precio_venta"]) - float(prod["precio_costo"])) * cant_salida, 2)
            registrar_venta_en_excel(prod["codigo"], prod["nombre"], cant_salida, prod["precio_venta"], gan_v)
            guardar_inventario(inventario)
            actualizar_vista()
            mostrar_mensaje(f"Venta registrada. Ganancia: L. {gan_v:.2f}")
            page.dialog.open = False
            page.update()

        page.dialog = ft.AlertDialog(
            title=ft.Text(f"Vender: {prod['nombre']}"),
            content=ft.Column([ft.Text(f"Disponible: {prod['cantidad']}"), txt_cant_venta], tight=True),
            actions=[
                ft.ElevatedButton("Confirmar Venta", on_click=confirmar_venta),
                ft.TextButton("Cancelar", on_click=lambda _: setattr(page.dialog, 'open', False) or page.update())
            ]
        )
        page.dialog.open = True
        page.update()

    def eliminar_producto(prod):
        def confirmar(e):
            inventario.remove(prod)
            guardar_inventario(inventario)
            actualizar_vista()
            page.dialog.open = False
            mostrar_mensaje("Producto eliminado.", ft.Colors.RED_400)

        page.dialog = ft.AlertDialog(
            title=ft.Text("Confirmar eliminación"),
            content=ft.Text(f"¿Deseas eliminar '{prod.get('nombre')}'?"),
            actions=[
                ft.ElevatedButton("Eliminar", on_click=confirmar, bgcolor=ft.Colors.RED_600, color=ft.Colors.WHITE),
                ft.TextButton("Cancelar", on_click=lambda _: setattr(page.dialog, 'open', False) or page.update())
            ]
        )
        page.dialog.open = True
        page.update()

    def cargar_pantalla_principal():
        page.controls.clear()
        page.add(
            ft.Row([
                ft.Text("Donde La Sarita", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_800, expand=True),
                ft.IconButton(icon=ft.Icons.LOCK, tooltip="Bloquear App", on_click=lambda _: verificar_estado_seguridad()),
                ft.FloatingActionButton(icon=ft.Icons.ADD, mini=True, on_click=lambda _: abrir_dialogo_producto())
            ]),
            ft.Row([txt_buscar]),
            ft.Divider(height=10),
            lista_productos_view
        )
        actualizar_vista()

    # ------------------------------------------------------------------
    # PROTOCOLO DE SEGURIDAD (REGISTRO Y LOGIN)
    # ------------------------------------------------------------------
    def cargar_pantalla_registro():
        txt_reg_nombre = ft.TextField(label="Nombre Completo", autofocus=True)
        txt_reg_usuario = ft.TextField(label="Usuario (ej. sarita)")
        txt_reg_pin = ft.TextField(label="PIN de 6 dígitos", password=True, can_reveal_password=True, max_length=6, keyboard_type=ft.KeyboardType.NUMBER)

        def procesar_registro(e):
            nom = txt_reg_nombre.value.strip()
            usr = txt_reg_usuario.value.strip()
            pin = txt_reg_pin.value.strip()

            if not nom or not usr:
                mostrar_mensaje("Por favor ingresa tu nombre y usuario.", ft.Colors.RED_500)
                return

            if len(pin) != 6 or not pin.isdigit():
                mostrar_mensaje("El PIN debe ser exactamente de 6 números.", ft.Colors.RED_500)
                return

            guardar_usuario(nom, usr, pin)
            mostrar_mensaje("¡Registro exitoso! Ya puedes ingresar con tu PIN.", ft.Colors.GREEN_600)
            cargar_pantalla_login()

        page.controls.clear()
        page.add(
            ft.Container(
                alignment=ft.Alignment(0, 0),
                expand=True,
                content=ft.Card(
                    elevation=4,
                    content=ft.Container(
                        padding=25,
                        width=350,
                        content=ft.Column([
                            ft.Icon(ft.Icons.SHIELD_OUTLINED, size=50, color=ft.Colors.BLUE_800),
                            ft.Text("Registro de Acceso", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_900),
                            ft.Text("Crea tu cuenta administradora para proteger el inventario:", size=12, color=ft.Colors.GREY_700),
                            ft.Divider(height=10),
                            txt_reg_nombre,
                            txt_reg_usuario,
                            txt_reg_pin,
                            ft.Container(height=10),
                            ft.ElevatedButton("Crear Cuenta y Guardar PIN", on_click=procesar_registro, bgcolor=ft.Colors.BLUE_800, color=ft.Colors.WHITE, width=300)
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10)
                    )
                )
            )
        )
        page.update()

    def cargar_pantalla_login():
        txt_login_pin = ft.TextField(label="PIN de 6 dígitos", password=True, can_reveal_password=True, max_length=6, keyboard_type=ft.KeyboardType.NUMBER, autofocus=True)

        def validar_ingreso(e):
            pin_ingresado = txt_login_pin.value.strip()
            if verificar_pin(pin_ingresado):
                mostrar_mensaje("Acceso concedido.", ft.Colors.GREEN_600)
                cargar_pantalla_principal()
            else:
                mostrar_mensaje("PIN incorrecto. Inténtalo de nuevo.", ft.Colors.RED_500)
                txt_login_pin.value = ""
                page.update()

        txt_login_pin.on_submit = validar_ingreso

        page.controls.clear()
        page.add(
            ft.Container(
                alignment=ft.Alignment(0, 0),
                expand=True,
                content=ft.Card(
                    elevation=4,
                    content=ft.Container(
                        padding=25,
                        width=340,
                        content=ft.Column([
                            ft.Icon(ft.Icons.LOCK_OUTLINED, size=50, color=ft.Colors.BLUE_800),
                            ft.Text("Donde La Sarita", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_900),
                            ft.Text("Ingresa tu clave de 6 dígitos para continuar:", size=12, color=ft.Colors.GREY_700),
                            ft.Divider(height=10),
                            txt_login_pin,
                            ft.Container(height=10),
                            ft.ElevatedButton("Ingresar al Sistema", on_click=validar_ingreso, bgcolor=ft.Colors.GREEN_600, color=ft.Colors.WHITE, width=300)
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10)
                    )
                )
            )
        )
        page.update()

    def verificar_estado_seguridad():
        usuarios = cargar_usuarios()
        if not usuarios:
            cargar_pantalla_registro()
        else:
            cargar_pantalla_login()

    # Inicio del flujo con validación de seguridad
    verificar_estado_seguridad()

if __name__ == "__main__":
    ft.app(target=main)